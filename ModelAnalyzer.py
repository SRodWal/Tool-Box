# -*- coding: utf-8 -*-
"""
Created on Tue Apr 20 12:13:38 2021

@author: serw1
"""

import os
import pandas as pd
from openpyxl import load_workbook 
import datetime
import shutil

##### This function will append data to existing excel files
def append_df_to_excel(filename, df, sheet_name='Sheet1', startrow=None,
                       truncate_sheet=False, 
                       **to_excel_kwargs):
    """
    Append a DataFrame [df] to existing Excel file [filename]
    into [sheet_name] Sheet.
    If [filename] doesn't exist, then this function will create it.

    @param filename: File path or existing ExcelWriter
                     (Example: '/path/to/file.xlsx')
    @param df: DataFrame to save to workbook
    @param sheet_name: Name of sheet which will contain DataFrame.
                       (default: 'Sheet1')
    @param startrow: upper left cell row to dump data frame.
                     Per default (startrow=None) calculate the last row
                     in the existing DF and write to the next row...
    @param truncate_sheet: truncate (remove and recreate) [sheet_name]
                           before writing DataFrame to Excel file
    @param to_excel_kwargs: arguments which will be passed to `DataFrame.to_excel()`
                            [can be a dictionary]
    @return: None

    Usage examples:

    >>> append_df_to_excel('d:/temp/test.xlsx', df)

    >>> append_df_to_excel('d:/temp/test.xlsx', df, header=None, index=False)

    >>> append_df_to_excel('d:/temp/test.xlsx', df, sheet_name='Sheet2',
                           index=False)

    >>> append_df_to_excel('d:/temp/test.xlsx', df, sheet_name='Sheet2', 
                           index=False, startrow=25)

    (c) [MaxU](https://stackoverflow.com/users/5741205/maxu?tab=profile)
    """
    # Excel file doesn't exist - saving and exiting
    if not os.path.isfile(filename):
        df.to_excel(
            filename,
            sheet_name=sheet_name, 
            startrow=startrow if startrow is not None else 0, 
            **to_excel_kwargs)
        return
    
    # ignore [engine] parameter if it was passed
    if 'engine' in to_excel_kwargs:
        to_excel_kwargs.pop('engine')

    writer = pd.ExcelWriter(filename, engine='openpyxl', mode='a')

    # try to open an existing workbook
    writer.book = load_workbook(filename)
    
    # get the last row in the existing Excel sheet
    # if it was not specified explicitly
    if startrow is None and sheet_name in writer.book.sheetnames:
        startrow = writer.book[sheet_name].max_row

    # truncate sheet
    if truncate_sheet and sheet_name in writer.book.sheetnames:
        # index of [sheet_name] sheet
        idx = writer.book.sheetnames.index(sheet_name)
        # remove [sheet_name]
        writer.book.remove(writer.book.worksheets[idx])
        # create an empty sheet [sheet_name] using old index
        writer.book.create_sheet(sheet_name, idx)
    
    # copy existing sheets
    writer.sheets = {ws.title:ws for ws in writer.book.worksheets}

    if startrow is None:
        startrow = 0

    # write out the new sheet
    df.to_excel(writer, sheet_name, startrow=startrow, **to_excel_kwargs)

    # save the workbook
    writer.save()
#####

## Read files in directory - Reads all csv and list them for analysis
files = [f for f in os.listdir('.') if os.path.isfile(f)]
folderdir = "Desktop/Documents/Work/Proyectos - Jorge/EnergyAnalysis/FirstRun TEST"
filetype = ".csv"
mypath = folderdir+"/**/*"+filetype

names = []
for f in files:
    if filetype in f:
        names.append(f)
        
### Append solutions to results sheet - Edit formula #######
solnum = [2,5,4,8]
vec1 = [["G8","B3"],["F9","B4"]]
vec2 = [["N8","F3"],["N9","F4"],["N10","F5"],["N11","F6"],["N12","F7"]]
vec3 = [["N8","J3"],["N9","J4"],["N10","J5"],["N12","J6"]]
vec4 = [["K8","N3"],["K9","N4"],["K10","N5"],["K11","N6"],["K12","N7"],["K15","N8"],["K16","N9"],["K17","N10"]]
solvec = [vec1,vec2,vec3,vec4]
        
for file_name in names:
    stname = file_name[0:len(file_name)-14].replace("Hourly data, ","")
    stname = stname.replace(" 2019-2022,","")
    stname = stname.replace(" CZ6,","")

    output_name = file_name[0:len(file_name)-4]+".xlsx"
    tempname = "TESTER_temp.xlsx"

    df = pd.read_csv(file_name)
    df_raw = df
    ### Read template and creat new file
    tabtype = ["Hourly","Cooling","Heating","HPWH","Results"]
    sheetnames = ["Raw", stname+", Hourly", stname+", Cooling", stname+", Heating", stname+", HPWH", stname+", Results"]
    shutil.copy(tempname,output_name)
    ss= load_workbook(output_name)
    for st, name in zip(ss.worksheets[1:6], tabtype):
        st.title = stname+", "+name
    sheet = ss[sheetnames[-1]]
    for tab, vec in zip(sheetnames[1:5],solvec):
        for sol in vec:
          sheet[sol[1]] = "='"+tab+"'!"+sol[0]
            
    
    
    ss.save(output_name)
    
    
    
    ######### Edit Date & Time    
    time_str = df.columns[0]
    timevec = df[time_str]
    time_s = time_str.strip("Hours since")
    t0 = time_s[0:5]
    m0 = time_s[6:9]
    d0 =  time_s[9:len(time_str)]

    def monthToNum(shortMonth):
        
        return {
                'jan' : 1,
                'feb' : 2,
                'mar' : 3,
                'apr' : 4,
                'may' : 5,
                'jun' : 6,
                'jul' : 7,
                'aug' : 8,
                'sep' : 9, 
                'oct' : 10,
                'nov' : 11,
                'dec' : 12
                }[shortMonth]

    #Suponemos que las mediciones empiezan a las 12AM
    #Usamos un formato con AM/PM
    fmt = "%m/%d/%Y %H:%M %p"
    time0 =  datetime.datetime(2021, monthToNum(m0.lower()),int(d0))
    timevec = [time0]
    strtimevec = [time0.strftime(fmt)]    
    k=0
    for i in range(len(df[time_str])-1):
        timevec.append(timevec[-1]+datetime.timedelta(hours = 1))
        strtimevec.append(timevec[-1].strftime(fmt))
  
    df[time_str] = timevec
    df.rename(columns = {df.columns[0] : "Date & Time"}, inplace = True)
    dfnet = df[df.columns[14]]
    netname = df.columns[14]
    df.drop(netname, axis = 1, inplace = True)
    df.insert(2, netname, dfnet)
    
    #################
    coolingdat = [1,4,7,15,18,19]
    heatingdat = [1,5,6,8,16,17]
    hpdat = [1,9,10,20]
    path = output_name
    append_df_to_excel(path, pd.read_csv(file_name), sheet_name = "Raw", index = False, startrow = 0)
    append_df_to_excel(path, df[[df.columns[0],df.columns[1],df.columns[2]]],sheet_name = sheetnames[1], index = False, startrow = 5)
    append_df_to_excel(path, df[[df.columns[0],df.columns[3],df.columns[6],df.columns[14],df.columns[17],df.columns[18]]],sheet_name = sheetnames[2], index = False, startrow = 5)
    append_df_to_excel(path, df[[df.columns[0],df.columns[4],df.columns[5],df.columns[7],df.columns[15],df.columns[16]]],sheet_name = sheetnames[3], index = False, startrow = 5)
    append_df_to_excel(path, df[[df.columns[0],df.columns[8],df.columns[9],df.columns[19]]],sheet_name = sheetnames[4], index = False, startrow = 5)
    